import io

import pandas as pd
from openpyxl import load_workbook

from promethee_core.table_export import export_table_to_excel, export_tables_to_excel


def test_export_table_writes_header_index_and_values():
    df = pd.DataFrame({"Rank": [1, 2], "Phi (net flow)": [0.4123, 0.1]}, index=["A1", "A2"])

    xlsx = export_table_to_excel(
        df, sheet_name="Flows", title="Net flows", index_label="Alternative",
        notes=[("Final ranking", "A1 > A2")],
    )
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb["Flows"]
    rows = list(ws.iter_rows(values_only=True))

    assert rows[0][0] == "Net flows"
    header_row = rows[2]
    assert header_row == ("Alternative", "Rank", "Phi (net flow)")
    assert rows[3] == ("A1", 1, 0.4123)
    assert rows[4] == ("A2", 2, 0.1)
    assert rows[-1] == ("Final ranking", "A1 > A2", None)


def test_export_table_formats_integers_and_floats_differently():
    df = pd.DataFrame({"Rank": [1], "Weight": [0.3333333]}, index=["A1"])
    xlsx = export_table_to_excel(df, index_label="Alternative")
    ws = load_workbook(io.BytesIO(xlsx)).active

    # header is row 1 here (no title), data starts row 2
    rank_cell = ws.cell(row=2, column=2)
    weight_cell = ws.cell(row=2, column=3)
    assert rank_cell.number_format == "0"
    assert weight_cell.number_format == "0.000"


def test_sheet_name_is_truncated_to_excels_limit():
    df = pd.DataFrame({"x": [1]}, index=["A1"])
    xlsx = export_table_to_excel(df, sheet_name="a" * 40)
    wb = load_workbook(io.BytesIO(xlsx))
    assert len(wb.sheetnames[0]) <= 31


def test_sheet_name_strips_characters_excel_forbids():
    # A criterion literally named "Price/kg" used to crash the export outright.
    df = pd.DataFrame({"x": [1]}, index=["A1"])
    xlsx = export_table_to_excel(df, sheet_name="Price/kg")
    wb = load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames[0] == "Price-kg"

    for bad in "[]:*?/\\":
        xlsx = export_table_to_excel(df, sheet_name=f"a{bad}b")
        wb = load_workbook(io.BytesIO(xlsx))
        assert bad not in wb.sheetnames[0]


def test_export_tables_sanitizes_and_dedupes_sheet_names():
    df = pd.DataFrame({"x": [1]}, index=["A1"])
    xlsx = export_tables_to_excel(
        [("Price/kg", df, None, None), ("Price:kg", df, None, None)], index_label="Alternative",
    )
    wb = load_workbook(io.BytesIO(xlsx))
    # both sanitize to "Price-kg"; the second must be deduped, not silently overwrite the first.
    assert len(wb.sheetnames) == 2
    assert len(set(wb.sheetnames)) == 2
    assert all("/" not in n and ":" not in n for n in wb.sheetnames)
