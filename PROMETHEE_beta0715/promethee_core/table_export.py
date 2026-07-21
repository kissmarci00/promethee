"""Excel export styled for publication: bold header row, thin borders, sized
columns, and clean decimal formatting. Used by the download buttons on the
Results page — either one table per file (``export_table_to_excel``) or
several tables as separate sheets of one file (``export_tables_to_excel``).
"""
from __future__ import annotations

import io

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2933", end_color="1F2933", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_SIDE = Side(style="thin", color="B0B0B0")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center")

_INVALID_SHEET_CHARS = set('[]:*?/\\')


def _safe_sheet_name(name: str) -> str:
    """Excel sheet titles can't contain any of ``[ ] : * ? / \\``, can't start
    or end with an apostrophe, can't be blank, and are capped at 31
    characters. Criterion/alternative names are free text (e.g. "Price/kg" or
    "Speed (km/h)"), so sanitize before ever using one as a sheet title."""
    cleaned = "".join("-" if ch in _INVALID_SHEET_CHARS else ch for ch in name).strip().strip("'")
    return (cleaned or "Sheet")[:31]


def _write_table(
    ws,
    df: pd.DataFrame,
    title: str | None = None,
    number_format: str = "0.000",
    index_label: str = "",
    notes: list[tuple[str, str]] | None = None,
) -> None:
    """Write a single styled table (bold header row, thin borders, sized
    columns) onto an existing, already-titled worksheet."""
    row = 1
    if title:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=13)
        row += 2

    header_row = row
    header_cell = ws.cell(row=header_row, column=1, value=index_label)
    header_cell.font, header_cell.fill, header_cell.border = HEADER_FONT, HEADER_FILL, THIN_BORDER
    for j, col in enumerate(df.columns, start=2):
        cell = ws.cell(row=header_row, column=j, value=str(col))
        cell.font, cell.fill, cell.border, cell.alignment = HEADER_FONT, HEADER_FILL, THIN_BORDER, CENTER

    for i, (idx, values) in enumerate(zip(df.index, df.to_numpy(dtype=object)), start=header_row + 1):
        label_cell = ws.cell(row=i, column=1, value=str(idx))
        label_cell.font, label_cell.border = Font(bold=True), THIN_BORDER
        for j, value in enumerate(values, start=2):
            is_int = isinstance(value, (int, np.integer)) and not isinstance(value, bool)
            is_float = isinstance(value, (float, np.floating))
            cell = ws.cell(row=i, column=j, value=(float(value) if is_float else int(value) if is_int else value))
            cell.border, cell.alignment = THIN_BORDER, CENTER
            if is_int:
                cell.number_format = "0"
            elif is_float:
                cell.number_format = number_format

    last_row = header_row + len(df.index)
    if notes:
        last_row += 1
        for label, value in notes:
            last_row += 1
            ws.cell(row=last_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=last_row, column=2, value=value)

    ws.column_dimensions["A"].width = max([len(index_label) + 2, 12] + [len(str(i)) + 2 for i in df.index])
    for j, col in enumerate(df.columns, start=2):
        widths = [len(str(col)) + 2] + [
            len(f"{v:.3f}") if isinstance(v, (float, np.floating)) else len(str(v)) for v in df[col]
        ]
        ws.column_dimensions[get_column_letter(j)].width = max([10] + widths)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate


def export_table_to_excel(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    title: str | None = None,
    number_format: str = "0.000",
    index_label: str = "",
    notes: list[tuple[str, str]] | None = None,
) -> bytes:
    """Export a single DataFrame (row index + columns both treated as labels) to
    a styled .xlsx. ``notes`` are optional extra label/value rows appended below
    the table (used for e.g. the three PROMETHEE ranking orders)."""
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(sheet_name)
    _write_table(ws, df, title=title, number_format=number_format, index_label=index_label, notes=notes)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_tables_to_excel(
    tables: list[tuple[str, pd.DataFrame, str | None, list[tuple[str, str]] | None]],
    number_format: str = "0.000",
    index_label: str = "",
) -> bytes:
    """Export several DataFrames as separate sheets of a single styled .xlsx.
    Each entry in ``tables`` is (sheet_name, df, title, notes) — same meaning
    as the matching parameters of :func:`export_table_to_excel`."""
    wb = Workbook()
    wb.remove(wb.active)
    seen_names: dict[str, int] = {}
    for sheet_name, df, title, notes in tables:
        name = _safe_sheet_name(sheet_name)
        if name in seen_names:
            seen_names[name] += 1
            name = f"{name[:28]}_{seen_names[name]}"
        else:
            seen_names[name] = 0
        ws = wb.create_sheet(title=name)
        _write_table(ws, df, title=title, number_format=number_format, index_label=index_label, notes=notes)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
